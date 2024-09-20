import pandas as pd
import locale
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def respell_serbian_name(name):
    
    # Define a mapping of incorrect to correct letters
    correction_mapping = {
        'ic': 'ić',
        'dj': 'đ',
        'Dj': 'Đ'   # Replace 'đ' with 'dj'
    }
    
    # Iterate through the correction mapping and apply corrections
    for incorrect, correct in correction_mapping.items():
        name = name.replace(incorrect, correct)

    return name

def extract_initials(name):

    if name == "Uprava SPOJI":
        return name
    # Trenne den Namen in Worte
    words = name.split()

    # Hilfsfunktion, um den ersten Buchstaben zu finden, der kein Sonderzeichen ist
    def get_first_valid_char(word):
        for char in word:
            if char.isalnum():  # Alphanumerische Zeichen
                return char.upper()
        return ''  # Falls kein gültiges Zeichen gefunden wird

    # Falls der Name aus einem Wort besteht (z.B. Firmenname), nimm die ersten zwei Buchstaben
    if len(words) == 1:
        return words[0][:2].upper() + "."

    # Falls der Name aus mehr als einem Wort besteht, finde die Initialen
    first_initial = get_first_valid_char(words[0])

    # Zweites Wort: Sonderzeichen überspringen oder nächstes Wort verwenden
    second_initial = ""
    for word in words[1:]:
        second_initial = get_first_valid_char(word)
        if second_initial:  # Initiale gefunden
            break

    # Falls kein zweites Wort oder gültiges Zeichen gefunden wurde, nur das erste Initial anzeigen
    if not second_initial:
        return first_initial + "."

    return first_initial + "." + second_initial + "."

def map_country(code):
    country_mapping = {
        'AT': 'Österreich',
        'DE': 'Deutschland',
        'FR': 'France',  # Add more mappings as needed
        'CH': 'Schweiz',
        'RS': 'Serbien'
    }
    # Use get method to handle cases where the code is not in the mapping
    return country_mapping.get(code, 'Österreich')

monate = [
    "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"
]

month_mapping = {
    "Januar": "Јануар",
    "Februar": "Фебруар",
    "März": "Март",
    "April": "Април",
    "Mai": "Мај",
    "Juni": "Јун",
    "Juli": "Јул",
    "August": "Август",
    "September": "Септембар",
    "Oktober": "Октобар",
    "November": "Новембар",
    "Dezember": "Децембар"
}

def convert_month_to_serbian_cyrillic(month_name):
    return month_mapping.get(month_name, month_name)  # Return the original name if not found


# Setze die Locale auf Deutsch (oder die entsprechende Locale deines Systems)
locale.setlocale(locale.LC_NUMERIC, 'de_DE.UTF-8')
def delocalize(string):
    try:
        return locale.atof(string)
    except ValueError:
        # Falls die obige Konvertierung fehlschlägt, versuche Punkte zu entfernen und konvertiere erneut
        return locale.atof(string.replace('.', ''))
year  = 2022
kontobericht = pd.read_csv(f"kontobericht{year}.csv", delimiter=',', encoding="utf-16")
kontobericht = kontobericht.rename(columns={'Valutadatum':'Datum','Buchungs-Details':'Сврха/ Zweck', 'Partnername':'Auftraggeber'})
kontobericht.drop(['Notiz', 'BIC/SWIFT', 'Partner Kontonummer', 'Bankleitzahl'], inplace=True, axis=1)
kontobericht['Betrag'] = kontobericht['Betrag'].apply(delocalize)

# Convert 'Datum' to datetime format
kontobericht['Datum'] = pd.to_datetime(kontobericht['Datum'], format='%d.%m.%Y')

# Filter rows with 'Betrag' > 0
kontobericht_eingang = kontobericht[kontobericht['Betrag'] > 0]
# Set 'Datum' as the index and sort the index
kontobericht_eingang = kontobericht_eingang.set_index('Datum').sort_index()
# Erstelle ein DataFrame für die Gesamtdaten
gesamt_daten = pd.DataFrame(columns=['Monat', 'Einnahmen'])

for i in range(1, 13):
    # Setze die maximale Anzahl der Tage basierend auf dem Monat
    max_day = 31 if i in [1, 3, 5, 7, 8, 10, 12] else 30 if i in [4, 6, 9, 11] else 28
    month_str = str(i).zfill(2)
    day_str = str(1).zfill(2)
    max_str = str(max_day).zfill(2)
    # Berechne die Summe für den Monat
    
    
    
    #date_str = f'{year}-' + month_str + '-' + day_str + ':' + f'{year}-' + month_str + '-' + max_str
    filtered_kontobericht_eingang = kontobericht_eingang[f'{year}-' + month_str + '-' + day_str : f'{year}-' + month_str + '-' + max_str]
    #filtered_kontobericht_eingang = filtered_kontobericht_eingang.rename(columns={'Datum':'Datum','Buchungs-Details':'Zweck', 'Partnername':'Auftraggeber'})
    #filtered_kontobericht_eingang['Datum'] = filtered_kontobericht_eingang['Datum'].dt.date
    # Berechne die Summe für den Monat
    monats_summe = filtered_kontobericht_eingang['Betrag'].sum()
    monat_string = convert_month_to_serbian_cyrillic(monate[i-1]) + ' / '+  monate[i-1]
    # Erstelle eine neue Zeile als DataFrame
    new_row = pd.DataFrame({'Monat': [monat_string], 'Einnahmen': [monats_summe]})

    # Verwende pd.concat, um die neue Zeile hinzuzufügen
    gesamt_daten = pd.concat([gesamt_daten, new_row], ignore_index=True)

    filtered_kontobericht_eingang.index = filtered_kontobericht_eingang.index.strftime('%d.%m.%Y')
    eingang_summe = filtered_kontobericht_eingang['Betrag'].sum()
    gesamt_daten_summe = gesamt_daten['Einnahmen'].sum()
    filtered_kontobericht_eingang.loc[:,'Betrag'] = filtered_kontobericht_eingang['Betrag'].apply(lambda x: '{:,.2f} €'.format(x))
    filtered_kontobericht_eingang.loc[:,'Сврха/ Zweck'] = "Донација / Spende"
    try:
        filtered_kontobericht_eingang.loc[:, 'Auftraggeber'].fillna('Uprava SPOJI', inplace=True)  # Fill empty cells with 'Uprava SPOJI'
        filtered_kontobericht_eingang.loc[:,'Auftraggeber'] = filtered_kontobericht_eingang['Auftraggeber'].apply(respell_serbian_name)
        filtered_kontobericht_eingang.loc[:,'Auftraggeber'] = filtered_kontobericht_eingang['Auftraggeber'].apply(extract_initials)
        filtered_kontobericht_eingang.loc[:,'Land'] = filtered_kontobericht_eingang['Partner IBAN'].str[:2].apply(map_country)
        filtered_kontobericht_eingang.loc[:, 'Land'].fillna('Österreich', inplace=True)
        filtered_kontobericht_eingang.drop(['Partner IBAN'], inplace=True, axis=1)
    except (AttributeError, IndexError) as e:
        print(f"Exception in month {month_str}: {e}")
        continue
    sheet_name=monate[i-1]
    serbian_cyrillic_month = convert_month_to_serbian_cyrillic(sheet_name)
    column_widths = {'A': 20, 'B': 30, 'C': 25, 'D': 15, 'E': 35}
 
    filtered_kontobericht_eingang = filtered_kontobericht_eingang.rename(columns={'Auftraggeber':'Име / Name','Land':'Земља / Land', 'Betrag':'Износ / Betrag'})
    filtered_kontobericht_eingang = filtered_kontobericht_eingang[['Име / Name', 'Земља / Land', 'Износ / Betrag', 'Сврха/ Zweck']]
    

    with pd.ExcelWriter(f"Izvestaji/Donatori - Spender {year}.xlsx", engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
        thin = Side(border_style="thin", color="000000")  # Dünne schwarze Linie
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        filtered_kontobericht_eingang.to_excel(writer, sheet_name=sheet_name, index=True, startrow=2, startcol=0)

        
        # Greife auf das aktuelle Blatt zu
        sheet = writer.book[sheet_name]
        
        max_row = len(filtered_kontobericht_eingang) + 2  # Startrow is 2
        max_col = len(filtered_kontobericht_eingang.columns)  # Number of columns in the DataFrame

        # Define the border style
        thin = Side(border_style="thin", color="000000")  # Thin black line
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Apply the border to the entire DataFrame range
        for row in range(2, max_row + 2):  # From startrow 2 to the last row
            for col in range(max_col + 1):  # Loop through each column
                cell = sheet.cell(row=row, column=col + 1)  # Excel columns are 1-indexed
                cell.border = border
            
        #Füge einen Titel hinzu
        sheet['A1'] = f'Донатори {serbian_cyrillic_month} {year}.'
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A1'].font = Font(bold=True,color='FFFFFF', size=16,name='Georgia')
        sheet['A1'].fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')


        sheet.merge_cells('A1:E1')

        sheet['A2'] = f'Spender {sheet_name} {year}.'
        sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A2'].font = Font(bold=True,color='FFFFFF', size=16,name='Georgia')
        sheet['A2'].fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

        sheet.merge_cells('A2:E2')

        sum_einnahmen = round(eingang_summe, 2)
        
        sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)] = 'УКУПНО / GESAMT'
        sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].font = Font(bold=True, size=12,name='Georgia')
        sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        #sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].number_format = '#,##0.00_-'
        #sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].font = Font(u='double', color='000000')   
        sheet['D{}'.format(len(filtered_kontobericht_eingang) + 4)] = f"{sum_einnahmen} €"
        #sheet['D{}'.format(len(filtered_kontobericht_eingang) + 4)].fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        sheet['D{}'.format(len(filtered_kontobericht_eingang) + 4)].number_format = '#,##0.00_-'
        

        sheet['D{}'.format(len(filtered_kontobericht_eingang) + 4)].font = Font(u='single', color='000000', bold=True, size=12,name='Georgia') 

        row = len(filtered_kontobericht_eingang) + 4  
        sheet.merge_cells(f'A{row}:C{row}')

        

        for column, width in column_widths.items():
            sheet.column_dimensions[column].width = width
        for row_num, value in enumerate(filtered_kontobericht_eingang['Сврха/ Zweck'], start=2):  # Beginne bei 2, da wir bei startrow=1 beginnen
            cell = sheet.cell(row=row_num, column=2)
            cell.alignment = cell.alignment.copy(wrap_text=True)  # Aktiviere den Textumbruch
        
            # Schreibe die Gesamtdaten ins Excel-Blatt "GESAMT {year}"

with pd.ExcelWriter(f"Izvestaji/Donatori - Spender {year}.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    gesamt_sheet_name = f'GESAMT {year}'
    gesamt_daten.to_excel(writer, sheet_name=gesamt_sheet_name, startrow=2, index=False)

    # Formatieren des Gesamtblatts
    sheet = writer.book[gesamt_sheet_name]

    gesamt_daten.loc[:,'Einnahmen'] = gesamt_daten['Einnahmen'].apply(lambda x: '{:,.2f} €'.format(x))
    #Füge einen Titel hinzu
    sheet['A1'] = f'Донатори {year}.'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].font = Font(bold=True,color='FFFFFF', size=16,name='Georgia')
    sheet['A1'].fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')


    sheet.merge_cells('A1:B1')

    sheet['A2'] = f'Spender {year}.'
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A2'].font = Font(bold=True,color='FFFFFF', size=16,name='Georgia')
    sheet['A2'].fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

    sheet.merge_cells('A2:B2')

    sheet['A3'] = 'Месец / Monat'
    sheet['B3'] = 'Износ / Betrag (EUR)'
    sheet['A3'].font = Font(bold=True, size=12,name='Georgia')
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B3'].font = Font(bold=True, size=12,name='Georgia')
    sheet['B3'].alignment = Alignment(horizontal='center', vertical='center')

    sum_einnahmen = round(gesamt_daten_summe, 2)

    sheet['A{}'.format(len(gesamt_daten) + 4)] = 'УКУПНО / GESAMT'
    sheet['A{}'.format(len(gesamt_daten) + 4)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A{}'.format(len(gesamt_daten) + 4)].font = Font(bold=True, size=12,name='Georgia')
    sheet['A{}'.format(len(gesamt_daten) + 4)].fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    #sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].number_format = '#,##0.00_-'
    #sheet['A{}'.format(len(filtered_kontobericht_eingang) + 4)].font = Font(u='double', color='000000')   
    sheet['B{}'.format(len(gesamt_daten) + 4)] = f"{sum_einnahmen} €"
    sheet['B{}'.format(len(gesamt_daten) + 4)].alignment = Alignment(horizontal='right')
    #sheet['D{}'.format(len(filtered_kontobericht_eingang) + 4)].fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    sheet['B{}'.format(len(gesamt_daten) + 4)].number_format = '#,##0.00_-'


    sheet['B{}'.format(len(gesamt_daten) + 4)].font = Font(u='single', color='000000', bold=True, size=12,name='Georgia') 

    row = len(gesamt_daten) + 4  
    #sheet.merge_cells(f'A{row}:C{row}')

    # Formatieren der 'Einnahmen'-Spalte
    for idx in range(2, len(gesamt_daten) + 4):  # Start bei 2, weil die Daten bei row 2 beginnen
        cell = sheet.cell(row=idx, column=2)  # Spalte 2 ist 'Einnahmen'
        cell.number_format = '#,##0.00_-" €"'  

    # Hinzufügen von Rahmen
    for row in range(1, len(gesamt_daten) + 3):  # +2 für die Kopfzeile
        for col in range(2):  # 2 Spalten: Monat und Einnahmen
            cell = sheet.cell(row=row + 1, column=col + 1)  # Excel-Indizes sind 1-basiert
            cell.border = border  # Wenden Sie die vorher definierte Rahmenlinie an

    # Optional: Setze die Breite der Spalten
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 35
